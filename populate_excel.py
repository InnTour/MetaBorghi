#!/usr/bin/env python3
"""
Popola MetaBorghi_DataTemplate.xlsx con UN ESEMPIO per foglio (Lacedonia come riferimento).
Include il nuovo foglio Ristorazione.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

EXCEL_PATH = "MetaBorghi_DataTemplate.xlsx"

# ─────────────────────────────────────────────────
# MACRO-DESCRIZIONE LACEDONIA
# ─────────────────────────────────────────────────
LACEDONIA_MACRO = """Lacedonia: Borgo Autentico tra Storia Millenaria e Innovazione Digitale

Lacedonia si svela al visitatore come un borgo autentico dell'Alta Irpinia dove 13.000 anni di storia continuano a vivere tra vicoli medievali, palazzi nobiliari e tradizioni che il tempo non ha spezzato. Arroccato a 732 metri di altitudine, questo straordinario comune custodisce un'identità profonda che affonda le radici nell'antichità più remota e si proietta con determinazione verso il futuro.

Terra di Storia: dall'Antica Aquilonia ai Romani
Il patrimonio archeologico di Lacedonia affonda le radici nella civiltà sannitica, quando il territorio era parte dell'importante municipium romano di Aquilonia, centro strategico nell'organizzazione amministrativa e militare dell'Italia meridionale. I reperti romani rinvenuti nel corso dei secoli testimoniano la ricchezza e l'importanza di questo antico insediamento, che controllava i collegamenti tra la Campania interna e la Puglia. Le oltre 150 grotte tufacee abitate sin da 13.000 anni fa raccontano una continuità insediativa eccezionale, mentre il centro storico con le sue caratteristiche "strette" medievali che si intrecciano come un labirinto di pietra conserva l'atmosfera di un borgo dove il tempo rallenta, permettendo di riscoprire ritmi più umani e connessioni profonde.

Terra di Miracoli: San Gerardo Maiella e il Pozzo Prodigioso
Lacedonia è universalmente riconosciuta come Terra di Miracoli, legata indissolubilmente alla figura di San Gerardo Maiella, santo patrono delle mamme e dei bambini, venerato in tutto il mondo cattolico. Il celebre Pozzo del Miracolo rappresenta uno dei luoghi di pellegrinaggio più significativi dell'Irpinia, dove la tradizione popolare tramanda eventi prodigiosi legati al passaggio del santo. Il Palazzo Vescovile, che ospitò San Gerardo durante i suoi soggiorni a Lacedonia, è diventato meta di devozione per migliaia di fedeli che ogni anno giungono da ogni parte d'Italia e del mondo per rendere omaggio al santo e attingere alla spiritualità profonda che permea questi luoghi.

Terra di Cultura: Francesco De Sanctis e l'Illuminismo Meridionale
Lacedonia si distingue come Terra di Cultura grazie alla presenza dell'Istituto Magistrale fondato da Francesco De Sanctis nel 1878, uno dei primi esempi di scuola nell'Italia post-unificazione e simbolo dell'impegno per l'emancipazione culturale del Mezzogiorno. De Sanctis, eminente critico letterario, filosofo e uomo politico italiano, scelse la sua terra d'origine per realizzare un progetto educativo rivoluzionario: formare nuovi maestri elementari per diffondere il sapere nelle zone più periferiche del Sud Italia.

Patrimonio Religioso e Artistico di Eccellenza
La Concattedrale di Santa Maria Assunta domina il panorama con la sua imponente architettura barocca del XVII secolo. Il Museo Diocesano conserva preziose testimonianze d'arte sacra. La Chiesa di Santa Maria della Cancellata, edificata su un antico tempio romano dedicato alla dea Iside, testimonia la continuità millenaria delle presenze religiose. La leggenda popolare ha donato al territorio un'aura di mistero attraverso la celebre Casa del Diavolo, enigmatica struttura arroccata su una rupe tufacea.

Eccellenze Culturali Internazionali
Tra le attrazioni di eccellenza brilla il MAVI (Museo Antropologico Visivo Irpino), istituzione culturale unica a livello internazionale che custodisce 1.801 fotografie scattate nel 1957 dall'antropologo americano Frank Cancian. Questo straordinario archivio visivo documenta la vita quotidiana della comunità contadina irpina degli anni Cinquanta, offrendo un viaggio emozionante in un mondo rurale che rivive nella memoria collettiva.

Patrimonio Naturalistico e Tradizioni Millenarie
I Boschi dell'Origlio, con le loro sorgenti minerali dalle proprietà benefiche, sono un polmone verde di straordinaria ricchezza. Due alberi monumentali riconosciuti dalla Regione Campania, il Cerro del Tesoro e il Cerro del Drago, vegliano sul territorio. La Valle del Calaggio, teatro delle antiche guerre sannitiche, offre scenari di rara suggestione. La transumanza, riconosciuta Patrimonio UNESCO nel 2019, rappresenta un elemento identitario di eccezionale valore.

Eccellenze Enogastronomiche
Il territorio è rinomato per gli asparagi selvatici di Contrada Forna, lo zafferano irpino, la produzione casearia tradizionale con formaggi di straordinaria qualità e i dolci natalizi tramandati di generazione in generazione. Il territorio si inserisce nel circuito delle grandi produzioni DOCG dell'Irpinia.

Innovazione Digitale per il Futuro
Lacedonia si distingue oggi come borgo all'avanguardia nella trasformazione digitale dei territori rurali. Grazie alla piattaforma innovativa sviluppata da InnTour, oltre 150.000 mq del centro storico sono stati mappati con tecnologie di ultima generazione. Scansioni 3D, fotografie panoramiche a 360°, contenuti multimediali immersivi e avatar AI conversazionali permettono di esplorare ogni angolo in un'esperienza che fonde passato e futuro.

Scopri i sapori, vivi le tradizioni, respira la storia. Benvenuto a Lacedonia, dove ogni pietra racconta e ogni esperienza emoziona."""

# ─────────────────────────────────────────────────
# 1. AZIENDE — esempio Lacedonia (Caciocavalleria De D.)
# ─────────────────────────────────────────────────
AZIENDE_COLS = [
    "id", "slug", "name", "legal_name", "vat_number",
    "founding_year", "employees_count", "founder_name", "founder_quote",
    "borough_id", "address_full", "coordinates_lat", "coordinates_lng",
    "type", "tagline", "description_short", "description_long",
    "certifications", "awards", "tier",
    "contact_email", "contact_phone", "website_url",
    "social_instagram", "social_facebook",
    "main_video_url", "virtual_tour_url",
    "b2b_open_for_contact", "b2b_interests",
    "is_verified", "is_active",
]

azienda_esempio = {
    "id": "caciocavalleria",
    "slug": "caciocavalleria",
    "name": "Caciocavalleria De D.",
    "legal_name": "Caciocavalleria De D. S.a.s.",
    "vat_number": "IT07890123456",
    "founding_year": 1975,
    "employees_count": 10,
    "founder_name": "Paolo De Dominicis",
    "founder_quote": "Le nostre vacche Podoliche camminano libere. Il formaggio che producono non ha eguali.",
    "borough_id": "lacedonia",
    "address_full": "Contrada Masseria 5, 83046 Lacedonia (AV)",
    "coordinates_lat": 41.05,
    "coordinates_lng": 15.42,
    "type": "PRODUTTORE_FOOD",
    "tagline": "Il re dei formaggi irpini",
    "description_short": "Caciocavallo Podolico e formaggi di latte crudo dall'Irpinia.",
    "description_long": "La Caciocavalleria De D. è un caseificio artigianale specializzato nella produzione di Caciocavallo Podolico, ottenuto dal latte delle vacche Podoliche allevate al pascolo brado sui pascoli dell'Alta Irpinia. Questo formaggio raro, prodotto solo con latte di razza autoctona, viene stagionato per un minimo di 12 mesi nelle grotte naturali. Il risultato è un formaggio dal gusto intenso, leggermente piccante, con sentori di erbe selvatiche.",
    "certifications": "Slow Food, De.Co.",
    "awards": "2023 - Presidio Slow Food - Slow Food Italia",
    "tier": "PREMIUM",
    "contact_email": "info@caciocavalleriaded.it",
    "contact_phone": "+39 0827 85012",
    "website_url": "https://www.caciocavalleriaded.it",
    "social_instagram": "#",
    "social_facebook": "#",
    "main_video_url": "",
    "virtual_tour_url": "",
    "b2b_open_for_contact": "SI",
    "b2b_interests": "Distribuzione, Ristorazione, Export",
    "is_verified": "SI",
    "is_active": "SI",
}

# ─────────────────────────────────────────────────
# 2. ESPERIENZE — esempio Lacedonia (Tour Caseificio + MAVI)
# ─────────────────────────────────────────────────
ESPERIENZE_COLS = [
    "id", "slug", "title", "tagline", "category", "provider_id", "borough_id",
    "description_short", "description_long",
    "duration_minutes", "max_participants", "min_participants", "price_per_person",
    "difficulty_level", "languages_available",
    "includes", "excludes", "what_to_bring",
    "cancellation_policy", "accessibility_info", "seasonal_tags",
    "coordinates_lat", "coordinates_lng",
    "timeline_step_1_time", "timeline_step_1_title", "timeline_step_1_desc",
    "timeline_step_2_time", "timeline_step_2_title", "timeline_step_2_desc",
    "timeline_step_3_time", "timeline_step_3_title", "timeline_step_3_desc",
    "timeline_step_4_time", "timeline_step_4_title", "timeline_step_4_desc",
    "timeline_step_5_time", "timeline_step_5_title", "timeline_step_5_desc",
    "is_active",
]

esperienza_esempio = {
    "id": "tour-caseificio",
    "slug": "tour-caseificio-caciocavallo-podolico",
    "title": "Il Caciocavallo Podolico: Dal Pascolo alla Tavola",
    "tagline": "Un formaggio raro, una storia millenaria",
    "category": "GASTRONOMIA",
    "provider_id": "caciocavalleria",
    "borough_id": "lacedonia",
    "description_short": "Visita al caseificio artigianale con mungitura, caseificazione dal vivo e degustazione guidata di 5 stagionature.",
    "description_long": "Scopri il segreto del Caciocavallo Podolico, uno dei formaggi più rari d'Italia, nella Caciocavalleria De D. a Lacedonia. La giornata inizia con la visita ai pascoli dove le vacche Podoliche brucano libere. Si prosegue con la dimostrazione di caseificazione dal vivo. La visita si conclude nelle grotte di stagionatura in tufo. Degustazione finale di 5 stagionature diverse (3, 6, 12, 18 e 24 mesi), accompagnate da miele di castagno e confettura di fichi.",
    "duration_minutes": 180,
    "max_participants": 10,
    "min_participants": 2,
    "price_per_person": 55,
    "difficulty_level": "FACILE",
    "languages_available": "Italiano",
    "includes": "Visita pascoli|Dimostrazione caseificazione|Degustazione 5 stagionature|Miele e confettura|Forma di caciocavallo giovane omaggio",
    "excludes": "Trasporto|Pranzo",
    "what_to_bring": "Scarpe da campagna|Abbigliamento comodo",
    "cancellation_policy": "Cancellazione gratuita fino a 48 ore prima.",
    "accessibility_info": "Cantina accessibile con rampa. Visita pascoli su terreno non pavimentato.",
    "seasonal_tags": "primavera, estate, autunno",
    "coordinates_lat": 41.05,
    "coordinates_lng": 15.42,
    "timeline_step_1_time": "09:00",
    "timeline_step_1_title": "Pascoli",
    "timeline_step_1_desc": "Visita alle vacche Podoliche al pascolo",
    "timeline_step_2_time": "09:45",
    "timeline_step_2_title": "Caseificazione",
    "timeline_step_2_desc": "Dimostrazione di cagliatura e filatura a mano",
    "timeline_step_3_time": "10:45",
    "timeline_step_3_title": "Grotte di stagionatura",
    "timeline_step_3_desc": "Visita alle grotte in tufo con formaggi in affinamento",
    "timeline_step_4_time": "11:15",
    "timeline_step_4_title": "Degustazione",
    "timeline_step_4_desc": "5 stagionature con miele e confettura",
    "timeline_step_5_time": "",
    "timeline_step_5_title": "",
    "timeline_step_5_desc": "",
    "is_active": "SI",
}

# ─────────────────────────────────────────────────
# 3. ARTIGIANATO — esempio Lacedonia
# ─────────────────────────────────────────────────
ARTIGIANATO_COLS = [
    "id", "slug", "name", "artisan_id", "borough_id",
    "description_short", "description_long",
    "technique_description", "material_type",
    "price", "dimensions", "weight_grams",
    "is_unique_piece", "production_series_qty", "lead_time_days",
    "is_custom_order_available", "stock_qty",
    "custom_option_1_name", "custom_option_1_values", "custom_option_1_price_modifier",
    "custom_option_2_name", "custom_option_2_values", "custom_option_2_price_modifier",
    "process_step_1_title", "process_step_1_desc",
    "process_step_2_title", "process_step_2_desc",
    "process_step_3_title", "process_step_3_desc",
    "is_active",
]

artigianato_esempio = {
    "id": "cesto-vimini-lacedonia",
    "slug": "cesto-vimini-tradizionale-lacedonia",
    "name": "Cesto in Vimini per Raccolta Tradizionale",
    "artisan_id": "akudunniad",
    "borough_id": "lacedonia",
    "description_short": "Cesto intrecciato a mano in vimini locale, perfetto per raccolta e decorazione.",
    "description_long": "Cesto tradizionale intrecciato a mano con vimini raccolto lungo i fiumi dell'Irpinia. La tecnica di intreccio è quella tramandata dai contadini per la raccolta di castagne, olive e uva. Il manico rinforzato permette di trasportare carichi pesanti. Ogni cesto è unico per le variazioni naturali del vimini.",
    "technique_description": "Intreccio a mano con tecnica tradizionale, manico rinforzato",
    "material_type": "vimini",
    "price": 42,
    "dimensions": "diam. 35 cm x h 25 cm (con manico h 40 cm)",
    "weight_grams": 450,
    "is_unique_piece": "NO",
    "production_series_qty": 20,
    "lead_time_days": 12,
    "is_custom_order_available": "SI",
    "stock_qty": 15,
    "custom_option_1_name": "Dimensione",
    "custom_option_1_values": "Piccolo diam.25, Medio diam.35, Grande diam.45",
    "custom_option_1_price_modifier": 15,
    "custom_option_2_name": "",
    "custom_option_2_values": "",
    "custom_option_2_price_modifier": "",
    "process_step_1_title": "Raccolta vimini",
    "process_step_1_desc": "Vimini locale raccolto in inverno lungo i fiumi",
    "process_step_2_title": "Intreccio",
    "process_step_2_desc": "Intreccio a mano con tecnica tradizionale",
    "process_step_3_title": "",
    "process_step_3_desc": "",
    "is_active": "SI",
}

# ─────────────────────────────────────────────────
# 4. PRODOTTI FOOD — esempio Lacedonia (Caciocavallo Podolico)
# ─────────────────────────────────────────────────
PRODOTTI_COLS = [
    "id", "slug", "name", "producer_id", "borough_id", "category",
    "description_short", "description_long", "tagline", "pairing_suggestions",
    "price", "unit", "weight_grams", "shelf_life_days", "storage_instructions",
    "origin_protected", "allergens", "ingredients",
    "stock_qty", "min_order_qty", "is_shippable", "shipping_notes",
    "is_active", "is_featured",
]

prodotto_esempio = {
    "id": "caciocavallo-podolico-18m",
    "slug": "caciocavallo-podolico-18m",
    "name": "Caciocavallo Podolico 18 Mesi",
    "producer_id": "caciocavalleria",
    "borough_id": "lacedonia",
    "category": "FORMAGGI",
    "description_short": "Il re dei formaggi irpini: Caciocavallo Podolico stagionato 18 mesi in grotta di tufo.",
    "description_long": "Il Caciocavallo Podolico della Caciocavalleria De D. è un formaggio raro e prezioso, prodotto esclusivamente con latte crudo di vacche Podoliche allevate al pascolo brado sui pascoli dell'Alta Irpinia. La cagliatura tradizionale e la stagionatura di 18 mesi nelle grotte naturali in tufo conferiscono un sapore intenso e complesso: note di erbe selvatiche, nocciola, burro fuso e un leggero piccante che si attenua nel finale lungo e persistente.",
    "tagline": "Il re dei formaggi del Sud",
    "pairing_suggestions": "Miele di castagno, confettura di fichi, Taurasi DOCG",
    "price": 32,
    "unit": "pezzo (ca. 1.2 kg)",
    "weight_grams": 1200,
    "shelf_life_days": 180,
    "storage_instructions": "Conservare in luogo fresco e asciutto, avvolto in carta alimentare",
    "origin_protected": "Presidio Slow Food",
    "allergens": "Latte",
    "ingredients": "Latte crudo di vacca Podolica, caglio naturale, sale",
    "stock_qty": 25,
    "min_order_qty": 1,
    "is_shippable": "SI",
    "shipping_notes": "Spedizione in confezione isotermica",
    "is_active": "SI",
    "is_featured": "SI",
}

# ─────────────────────────────────────────────────
# 5. OSPITALITA — esempio Lacedonia
# ─────────────────────────────────────────────────
OSPITALITA_COLS = [
    "id", "slug", "name", "type", "provider_id", "borough_id",
    "address_full", "coordinates_lat", "coordinates_lng", "distance_center_km",
    "description_short", "description_long", "tagline",
    "rooms_count", "max_guests", "price_per_night_from", "stars_or_category",
    "check_in_time", "check_out_time", "min_stay_nights",
    "amenities", "accessibility", "languages_spoken", "cancellation_policy",
    "booking_email", "booking_phone", "booking_url",
    "main_video_url", "virtual_tour_url",
    "is_active", "is_featured",
]

ospitalita_esempio = {
    "id": "masseria-lacedonia",
    "slug": "masseria-santa-lucia-lacedonia",
    "name": "Masseria Santa Lucia",
    "type": "MASSERIA",
    "provider_id": "caciocavalleria",
    "borough_id": "lacedonia",
    "address_full": "Contrada Masseria 5, 83046 Lacedonia (AV)",
    "coordinates_lat": 41.05,
    "coordinates_lng": 15.42,
    "distance_center_km": 2.5,
    "description_short": "Masseria storica immersa nei pascoli dell'Alta Irpinia, a due passi dal caseificio del Caciocavallo Podolico.",
    "description_long": "Masseria Santa Lucia è un'antica masseria ristrutturata che offre un'esperienza di soggiorno autentica nel cuore dell'Alta Irpinia. Le camere, ricavate dalle antiche stalle e dai fienili, conservano le volte in pietra e i pavimenti in cotto originali. La colazione è a base di prodotti del caseificio e dell'orto aziendale. Posizione ideale per esplorare Lacedonia e i borghi circostanti.",
    "tagline": "Dormire tra storia e pascoli",
    "rooms_count": 5,
    "max_guests": 12,
    "price_per_night_from": 75,
    "stars_or_category": "Agriturismo 3 spighe",
    "check_in_time": "15:00",
    "check_out_time": "11:00",
    "min_stay_nights": 1,
    "amenities": "WiFi, Parcheggio, Colazione inclusa, Giardino, Animali ammessi",
    "accessibility": "Piano terra accessibile",
    "languages_spoken": "Italiano, English",
    "cancellation_policy": "Cancellazione gratuita fino a 48 ore prima",
    "booking_email": "booking@caciocavalleriaded.it",
    "booking_phone": "+39 0827 85012",
    "booking_url": "",
    "main_video_url": "",
    "virtual_tour_url": "",
    "is_active": "SI",
    "is_featured": "SI",
}

# ─────────────────────────────────────────────────
# 6. RISTORAZIONE (NUOVO) — esempio Lacedonia
# ─────────────────────────────────────────────────
RISTORAZIONE_COLS = [
    "id", "slug", "name", "type", "borough_id",
    "address_full", "coordinates_lat", "coordinates_lng",
    "description_short", "description_long", "tagline",
    "cuisine_type", "price_range",
    "seats_indoor", "seats_outdoor",
    "opening_hours", "closing_day",
    "specialties", "menu_highlights",
    "contact_email", "contact_phone", "website_url",
    "social_instagram", "social_facebook",
    "booking_url", "accepts_groups", "max_group_size",
    "b2b_open_for_contact", "b2b_interests",
    "is_active", "is_featured",
]

RISTORAZIONE_INSTRUCTIONS = {
    "id": "OBBLIGATORIO\nEs: ristorante-la-taverna-lacedonia",
    "slug": "OBBLIGATORIO\nURL-friendly. Es: taverna-irpinia-lacedonia",
    "name": "OBBLIGATORIO\nNome del locale",
    "type": "OBBLIGATORIO\nRISTORANTE / TRATTORIA / PIZZERIA / AGRITURISMO / ENOTECA / BAR / OSTERIA",
    "borough_id": "OBBLIGATORIO\nID del borgo. Es: lacedonia",
    "address_full": "OBBLIGATORIO\nIndirizzo completo. Es: Via Roma 1, 83046 Lacedonia (AV)",
    "coordinates_lat": "Facoltativo\nLatitudine GPS. Es: 41.0472",
    "coordinates_lng": "Facoltativo\nLongitudine GPS. Es: 15.4297",
    "description_short": "OBBLIGATORIO\nPer card (max 160 caratteri)",
    "description_long": "OBBLIGATORIO\nPagina dettaglio (max 800 caratteri)",
    "tagline": "OBBLIGATORIO\nSlogan breve (max 80 caratteri)",
    "cuisine_type": "OBBLIGATORIO\nEs: Tradizionale Irpina, Pizza, Pesce, Gourmet",
    "price_range": "OBBLIGATORIO\nBUDGET / MEDIO / ALTO / GOURMET",
    "seats_indoor": "Facoltativo\nPosti interni",
    "seats_outdoor": "Facoltativo\nPosti esterni/terrazza",
    "opening_hours": "Facoltativo\nEs: Mar-Dom 12:00-15:00, 19:00-23:00",
    "closing_day": "Facoltativo\nEs: Lunedi",
    "specialties": "OBBLIGATORIO\nPiatti tipici separati da virgola",
    "menu_highlights": "Facoltativo\nPiatti consigliati separati da |",
    "contact_email": "OBBLIGATORIO\nEmail del locale",
    "contact_phone": "Facoltativo\nTelefono. Es: +39 0827 12345",
    "website_url": "Facoltativo\nURL sito web",
    "social_instagram": "Facoltativo\nURL o @handle Instagram",
    "social_facebook": "Facoltativo\nURL pagina Facebook",
    "booking_url": "Facoltativo\nURL prenotazione (TheFork, Google, proprio sito)",
    "accepts_groups": "Facoltativo\nSI / NO",
    "max_group_size": "Facoltativo\nMax persone per gruppi/eventi",
    "b2b_open_for_contact": "Facoltativo\nSI / NO — disponibile per contatti B2B",
    "b2b_interests": "Facoltativo\nEs: Forniture locali, Catering, Eventi, Gruppi turistici",
    "is_active": "OBBLIGATORIO\nSI / NO",
    "is_featured": "Facoltativo\nSI / NO — metti in evidenza",
}

ristorazione_esempio = {
    "id": "trattoria-del-borgo-lacedonia",
    "slug": "trattoria-del-borgo-lacedonia",
    "name": "Trattoria del Borgo",
    "type": "TRATTORIA",
    "borough_id": "lacedonia",
    "address_full": "Via Roma 23, 83046 Lacedonia (AV)",
    "coordinates_lat": 41.0472,
    "coordinates_lng": 15.4297,
    "description_short": "Cucina tradizionale irpina nel cuore del centro storico di Lacedonia, con piatti della nonna e prodotti a km zero.",
    "description_long": "La Trattoria del Borgo è il luogo dove la cucina irpina vive nella sua forma più autentica. Situata nel centro storico di Lacedonia, tra le antiche \"strette\" medievali, propone un menu che cambia con le stagioni: fusilli al ferretto con ragù di castrato, zuppe di legumi con cotiche, caciocavallo podolico alla piastra, e i dolci natalizi della tradizione lacedoniese. Tutti gli ingredienti provengono da produttori locali e dall'orto di famiglia. In estate si mangia nella terrazza panoramica con vista sulla Valle del Calaggio.",
    "tagline": "I sapori autentici dell'Irpinia a tavola",
    "cuisine_type": "Tradizionale Irpina",
    "price_range": "MEDIO",
    "seats_indoor": 40,
    "seats_outdoor": 20,
    "opening_hours": "Mar-Dom 12:00-15:00, 19:00-22:30",
    "closing_day": "Lunedi",
    "specialties": "Fusilli al ferretto con ragù di castrato, Caciocavallo Podolico alla piastra, Zuppa di legumi irpini, Dolci natalizi lacedoniesi",
    "menu_highlights": "Antipasto del pastore (salumi, formaggi, sottoli)|Fusilli al ragù di castrato|Agnello alla brace con patate|Torta di castagne",
    "contact_email": "info@trattoriadelborgo.it",
    "contact_phone": "+39 0827 85100",
    "website_url": "",
    "social_instagram": "@trattoriadelborgo",
    "social_facebook": "",
    "booking_url": "",
    "accepts_groups": "SI",
    "max_group_size": 30,
    "b2b_open_for_contact": "SI",
    "b2b_interests": "Forniture locali, Gruppi turistici, Catering eventi",
    "is_active": "SI",
    "is_featured": "SI",
}

# ─────────────────────────────────────────────────
# 7. BORGHI — esempio Lacedonia con macro-descrizione
# ─────────────────────────────────────────────────
BORGHI_COLS = [
    "id", "name", "tagline",
    "description_short", "description_long",
    "highlight_1", "highlight_2", "highlight_3",
    "notable_experiences",
    "altitude_m", "area_km2", "population", "province", "region",
    "companies_count", "experiences_count", "accommodations_count",
    "main_video_url", "virtual_tour_url",
]

borgo_esempio = {
    "id": "lacedonia",
    "name": "Lacedonia",
    "tagline": "Terra di Storia, Terra di Miracoli, Terra di Cultura",
    "description_short": "Borgo autentico dell'Alta Irpinia dove 13.000 anni di storia vivono tra vicoli medievali, palazzi nobiliari e tradizioni millenarie. Arroccato a 732 m, Lacedonia custodisce un patrimonio unico tra archeologia, fede e innovazione digitale.",
    "description_long": LACEDONIA_MACRO,
    "highlight_1": "MAVI — Museo Antropologico Visivo Irpino (1.801 foto di Frank Cancian, 1957)",
    "highlight_2": "Terra di Miracoli — San Gerardo Maiella e il Pozzo Prodigioso",
    "highlight_3": "Innovazione Digitale — 150.000 mq mappati in 3D da InnTour",
    "notable_experiences": "MAVI, Museo Diocesano, Concattedrale S. Maria Assunta, Pozzo del Miracolo, Casa del Diavolo, Grotte Tufacee, Boschi dell'Origlio, Valle del Calaggio, Cammino dei Tratturi, Tour Caseificio Podolico",
    "altitude_m": 732,
    "area_km2": 79.5,
    "population": 2300,
    "province": "Avellino",
    "region": "Campania",
    "companies_count": 1,
    "experiences_count": 3,
    "accommodations_count": 1,
    "main_video_url": "https://www.youtube.com/embed/XhB4SonU7Pw?autoplay=1&mute=1&loop=1&playlist=XhB4SonU7Pw",
    "virtual_tour_url": "https://my.treedis.com/tour/lacedonia-vrcerogn",
}


# ═══════════════════════════════════════════════════
# FUNZIONI DI SCRITTURA
# ═══════════════════════════════════════════════════

# Stili condivisi
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E79")
INSTR_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
DATA_FONT = Font(name="Calibri", size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def write_sheet(ws, cols, data_dict, start_row=4):
    """Scrive una riga dati nel foglio a partire da start_row."""
    for c_idx, col_key in enumerate(cols):
        col_num = c_idx + 1
        value = data_dict.get(col_key, "")
        cell = ws.cell(row=start_row, column=col_num, value=value)
        cell.font = DATA_FONT
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER


def create_ristorazione_sheet(wb):
    """Crea il foglio Ristorazione da zero con header, istruzioni e dati esempio."""
    # Inserisci prima di Borghi (o alla fine se non esiste)
    if "Ristorazione" in wb.sheetnames:
        del wb["Ristorazione"]

    # Trova la posizione giusta (dopo Ospitalità, prima di Borghi)
    target_idx = len(wb.sheetnames)
    for i, name in enumerate(wb.sheetnames):
        if name == "Borghi":
            target_idx = i
            break
    ws = wb.create_sheet("Ristorazione", target_idx)

    # Sezioni header (riga 1)
    sections = {
        1: "🍽️ ANAGRAFICA",
        5: "📍 LOCALIZZAZIONE",
        9: "📝 DESCRIZIONI",
        12: "⚙️ DETTAGLI",
        16: "📞 CONTATTI / SOCIAL",
        24: "🤝 B2B",
        27: "📊 STATO",
    }
    for col, title in sections.items():
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = WRAP_ALIGN
    # Fill section bg for all cols
    for c in range(1, len(RISTORAZIONE_COLS) + 1):
        ws.cell(row=1, column=c).fill = SECTION_FILL
        ws.cell(row=1, column=c).border = THIN_BORDER

    # Header (riga 2)
    for c_idx, col_key in enumerate(RISTORAZIONE_COLS):
        col_num = c_idx + 1
        cell = ws.cell(row=2, column=col_num, value=col_key)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER

    # Istruzioni (riga 3)
    for c_idx, col_key in enumerate(RISTORAZIONE_COLS):
        col_num = c_idx + 1
        instr = RISTORAZIONE_INSTRUCTIONS.get(col_key, "Facoltativo")
        # Marca obbligatorio/facoltativo
        if instr.startswith("OBBLIGATORIO"):
            prefix = "✦ "
        else:
            prefix = "○ "
        cell = ws.cell(row=3, column=col_num, value=prefix + instr)
        cell.font = INSTR_FONT
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER

    # Dati esempio (riga 4)
    write_sheet(ws, RISTORAZIONE_COLS, ristorazione_esempio, start_row=4)

    # Larghezze colonne
    for c in range(1, len(RISTORAZIONE_COLS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 22

    return ws


# ═══════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════

def populate_all():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # --- 1. AZIENDE ---
    ws = wb["Aziende"]
    # Pulisci dati vecchi (da riga 4 in poi)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    write_sheet(ws, AZIENDE_COLS, azienda_esempio, start_row=4)
    print("  Aziende: 1 esempio (Caciocavalleria De D. — Lacedonia)")

    # --- 2. ESPERIENZE ---
    ws = wb["Esperienze"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    write_sheet(ws, ESPERIENZE_COLS, esperienza_esempio, start_row=4)
    print("  Esperienze: 1 esempio (Tour Caseificio Podolico — Lacedonia)")

    # --- 3. ARTIGIANATO ---
    ws = wb["Artigianato"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    write_sheet(ws, ARTIGIANATO_COLS, artigianato_esempio, start_row=4)
    print("  Artigianato: 1 esempio (Cesto in Vimini — Lacedonia)")

    # --- 4. PRODOTTI FOOD ---
    ws = wb["Prodotti Food"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    write_sheet(ws, PRODOTTI_COLS, prodotto_esempio, start_row=4)
    print("  Prodotti Food: 1 esempio (Caciocavallo Podolico 18m — Lacedonia)")

    # --- 5. OSPITALITA ---
    ws = wb["Ospitalità"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    write_sheet(ws, OSPITALITA_COLS, ospitalita_esempio, start_row=4)
    print("  Ospitalità: 1 esempio (Masseria Santa Lucia — Lacedonia)")

    # --- 6. RISTORAZIONE (NUOVO FOGLIO) ---
    create_ristorazione_sheet(wb)
    print("  Ristorazione: NUOVO FOGLIO creato + 1 esempio (Trattoria del Borgo — Lacedonia)")

    # --- 7. BORGHI ---
    ws = wb["Borghi"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    write_sheet(ws, BORGHI_COLS, borgo_esempio, start_row=4)
    print("  Borghi: 1 esempio (Lacedonia — macro-descrizione completa)")

    # --- SALVA ---
    wb.save(EXCEL_PATH)
    print(f"\nFile salvato: {EXCEL_PATH}")
    print("Tutti i fogli sono stati popolati con il formato Lacedonia.")


if __name__ == "__main__":
    print("Popolo MetaBorghi_DataTemplate.xlsx con esempio Lacedonia...\n")
    populate_all()
